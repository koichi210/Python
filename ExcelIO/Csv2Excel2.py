# -*- coding: utf-8 -*-
#WriteExcelCell

import openpyxl

TemplateExcelName = './Sample/Template.xlsx'
WriteExcelName = './Sample/Template_new.xlsx'
SampleData = 'Test'


# Bookを開く
WorkBookDest = openpyxl.load_workbook(TemplateExcelName)

# Sheetを選択
WorkSpaceDest = WorkBookDest.active

# Cellの値を更新
print(WorkSpaceDest.cell(row=1, column=1).value)
WorkSpaceDest.cell(row=1, column=1).value = SampleData
print(WorkSpaceDest.cell(row=1, column=1).value)

WorkBookDest.save(WriteExcelName)
